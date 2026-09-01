"""
Hojas de cálculo como datos, no como texto.

El problema que resuelve este módulo, dicho con el caso real que lo motivó:
preguntado por cuántas cuentas tiene un balance de 85, el asistente
respondía "3" -- las que salieron en los cuatro fragmentos de texto que vio
de unos setecientos. No era falta de contexto. Contar, listar y sumar son
operaciones sobre DATOS, y un modelo que ve una muestra no puede hacerlas
por mucho que se le agrande la ventana.

Así que la hoja se guarda con su forma -- encabezados y filas -- y esas
preguntas se **calculan**. Al modelo le llega el resultado ya calculado,
para que lo redacte. Es la misma frontera que rige analitica-puc: el motor
calcula, la IA escribe, y nunca al revés.

Lo que se calcula aquí es exacto y reproducible. Si una hoja supera el tope
de filas, se dice: un conteo sobre datos truncados es un piso, no un total,
y presentarlo como total sería exactamente el error que este módulo existe
para eliminar.
"""
from __future__ import annotations

import json
import os
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import db

# Tope de filas por hoja. 50.000 filas de un balance son ~15 MB en JSON:
# manejable. Más allá se guarda lo que cabe y se declara la truncadura.
TOPE_FILAS = int(os.getenv("TABLA_TOPE_FILAS", "50000"))

# Cuántos valores distintos se enumeran de una columna. Ochenta y cinco
# códigos de cuenta caben de sobra; una columna de descripciones libres con
# miles de valores distintos no aporta enumerada, y se resume.
TOPE_DISTINTOS = int(os.getenv("TABLA_TOPE_DISTINTOS", "300"))


# =====================================================================
# LECTURA
# =====================================================================

def _celda(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _numero(s: str) -> Decimal | None:
    """El valor numérico de una celda, o None si no lo es.

    Acepta la escritura colombiana (1.234.567,89) y la inglesa
    (1,234,567.89) decidiendo por cuál separador va más a la derecha, que
    es el decimal. Sin esto, una columna de saldos se sumaría mal o no se
    sumaría, que en un papel de auditoría es igual de inútil.
    """
    s = (s or "").strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1].strip()
    # Con letras no es una cifra. Sin esta guarda, quitar los caracteres no
    # numéricos convertía "Cuenta 105" en 105 y una columna de nombres se
    # sumaba como si fueran importes: un total perfectamente formado y sin
    # ningún significado.
    if re.search(r"[^\W\d_]", s, re.UNICODE):
        return None
    s = re.sub(r"[^\d.,+-]", "", s)
    if not s or not re.search(r"\d", s):
        return None
    if "." in s and "," in s:
        s = (s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".")
             else s.replace(",", ""))
    elif "," in s:
        s = s.replace(",", ".") if len(s) - s.rfind(",") - 1 <= 2 else s.replace(",", "")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    return -d if neg else d


def _encabezado(filas: list[list[str]]) -> int:
    """En qué fila están los encabezados.

    Los exportes de un ERP traen título, NIT y fechas antes de la tabla. Se
    toma la primera fila con al menos tres celdas no vacías y sin números:
    un encabezado son nombres, una fila de datos casi siempre trae cifras.
    """
    for i, f in enumerate(filas[:40]):
        llenas = [c for c in f if c]
        if len(llenas) >= 3 and not any(_numero(c) is not None for c in llenas):
            return i
    return 0


def leer_excel(ruta: Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(str(ruta), read_only=True, data_only=True)
    hojas = []
    for h in wb.worksheets:
        crudas = [[_celda(c) for c in fila]
                  for fila in h.iter_rows(values_only=True)]
        crudas = [f for f in crudas if any(f)]
        if len(crudas) < 2:
            continue
        i = _encabezado(crudas)
        columnas = [c or f"columna_{j + 1}"
                    for j, c in enumerate(crudas[i])]
        cuerpo = crudas[i + 1:]
        truncada = len(cuerpo) > TOPE_FILAS
        hojas.append({"hoja": h.title, "columnas": columnas,
                      "filas": cuerpo[:TOPE_FILAS],
                      "n_filas": len(cuerpo), "truncada": truncada})
    wb.close()
    return hojas


def leer_csv(ruta: Path) -> list[dict]:
    import csv
    for cod in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            texto = ruta.read_text(encoding=cod)
            break
        except UnicodeDecodeError:
            continue
    else:
        return []
    # El separador se detecta del contenido: en Colombia el punto y coma es
    # tan común como la coma, porque la coma es el decimal.
    muestra = "\n".join(texto.splitlines()[:20])
    try:
        sep = csv.Sniffer().sniff(muestra, delimiters=",;\t|").delimiter
    except csv.Error:
        sep = ";" if muestra.count(";") > muestra.count(",") else ","
    filas = [[_celda(c) for c in f]
             for f in csv.reader(texto.splitlines(), delimiter=sep)]
    filas = [f for f in filas if any(f)]
    if len(filas) < 2:
        return []
    i = _encabezado(filas)
    columnas = [c or f"columna_{j + 1}" for j, c in enumerate(filas[i])]
    cuerpo = filas[i + 1:]
    return [{"hoja": ruta.stem, "columnas": columnas,
             "filas": cuerpo[:TOPE_FILAS], "n_filas": len(cuerpo),
             "truncada": len(cuerpo) > TOPE_FILAS}]


def leer(ruta: Path, tipo: str) -> list[dict]:
    if tipo == "Excel":
        return leer_excel(ruta)
    if tipo == "CSV":
        return leer_csv(ruta)
    return []


# =====================================================================
# PERSISTENCIA
# =====================================================================

def guardar(documento_id: str, hojas: list[dict]) -> int:
    with db.conn() as c:
        c.execute("DELETE FROM app.tabla WHERE documento_id=%s", (documento_id,))
        for h in hojas:
            c.execute(
                """INSERT INTO app.tabla
                     (documento_id, hoja, columnas, filas, n_filas, truncada)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (documento_id, h["hoja"], json.dumps(h["columnas"]),
                 json.dumps(h["filas"]), h["n_filas"], h["truncada"]))
        c.commit()
    return len(hojas)


def tablas_de(doc_ids: list[str]) -> list[dict]:
    if not doc_ids:
        return []
    return db.varios(
        """SELECT t.*, d.nombre AS documento
             FROM app.tabla t
             JOIN app.documento d ON d.id = t.documento_id
            WHERE t.documento_id = ANY(%s)
            ORDER BY d.nombre, t.hoja""",
        (doc_ids,))


# =====================================================================
# LO QUE SE CALCULA
# =====================================================================

# Nombres que delatan un identificador. Un código de cuenta, un NIT o un
# número de factura son números que NO se suman: sumarlos da una cifra
# perfectamente formada y sin ningún significado.
NOMBRE_IDENTIFICADOR = re.compile(
    r"c[oó]digo|cuenta|nit|c[eé]dula|cedula|documento|factura|comprobante|"
    r"n[uú]mero|numero|" + chr(92) + r"bnum" + chr(92) + r"b|" + chr(92) + r"bid" + chr(92) + r"b|referencia|tercero",
    re.IGNORECASE)


def _es_identificador(nombre: str, llenos: list[str]) -> bool:
    """¿Esta columna numérica es en realidad un identificador?

    Dos señales, y basta una. La del NOMBRE es la fuerte: si la columna se
    llama "Código" o "NIT", no hay más que discutir. La ESTRUCTURAL salva el
    caso del encabezado raro o en otro idioma: dígitos puros, todos del
    mismo largo y todos distintos es la forma de un consecutivo, no la de
    una columna de importes -- los importes repiten y varían de largo.
    """
    if NOMBRE_IDENTIFICADOR.search(nombre or ""):
        return True
    if not llenos or not all(v.isdigit() for v in llenos):
        return False
    return len({len(v) for v in llenos}) == 1 and len(set(llenos)) == len(llenos)


def perfil_columna(columnas: list[str], filas: list[list[str]],
                   j: int) -> dict:
    """Qué hay en una columna: cuántos valores, cuáles, y su total si es
    numérica. Todo contado sobre las filas completas, no sobre una muestra."""
    valores = [f[j].strip() if j < len(f) else "" for f in filas]
    llenos = [v for v in valores if v]
    numeros = [_numero(v) for v in llenos]
    numericos = [n for n in numeros if n is not None]
    # Una columna es numérica si la gran mayoría de sus celdas lo son; unas
    # pocas de texto suelen ser subtotales o notas intercaladas.
    nombre = columnas[j] if j < len(columnas) else f"columna_{j+1}"
    es_num = bool(llenos) and len(numericos) >= len(llenos) * 0.8
    identificador = es_num and _es_identificador(nombre, llenos)
    if identificador:
        es_num = False        # se enumera, no se suma

    perfil = {"columna": nombre, "con_valor": len(llenos),
              "vacias": len(valores) - len(llenos), "numerica": es_num,
              "identificador": identificador}

    if es_num:
        perfil.update(
            suma=str(sum(numericos)),
            minimo=str(min(numericos)), maximo=str(max(numericos)))
    else:
        distintos = sorted(set(llenos))
        perfil["distintos"] = len(distintos)
        if len(distintos) <= TOPE_DISTINTOS:
            perfil["valores"] = distintos
        else:
            perfil["valores"] = distintos[:20]
            perfil["valores_truncados"] = True
    return perfil


def ficha(doc_ids: list[str]) -> str:
    """El bloque que se le pone al modelo: la forma real de cada tabla y sus
    cifras ya calculadas.

    Deliberadamente NO son fragmentos. Aquí no hay muestreo: el conteo de
    filas es el conteo de filas, y la suma es la suma de todas.
    """
    partes = []
    for t in tablas_de(doc_ids):
        cols, filas = t["columnas"], t["filas"]
        cab = (f'TABLA · {t["documento"]} · hoja "{t["hoja"]}"\n'
               f'{t["n_filas"]} filas de datos, {len(cols)} columnas.')
        if t["truncada"]:
            cab += (f'\nADVERTENCIA: solo se guardaron {len(filas)} de las '
                    f'{t["n_filas"]} filas. Todo conteo de abajo es un MÍNIMO, '
                    'no un total.')
        lineas = [cab, "Columnas y su contenido, calculado sobre TODAS las filas:"]

        for j in range(len(cols)):
            p = perfil_columna(cols, filas, j)
            if p["numerica"]:
                lineas.append(
                    f'· {p["columna"]}: numérica · {p["con_valor"]} valores · '
                    f'suma {p["suma"]} · mínimo {p["minimo"]} · máximo {p["maximo"]}')
            elif p.get("valores_truncados"):
                lineas.append(
                    f'· {p["columna"]}: texto · {p["distintos"]} valores '
                    f'distintos (demasiados para listar; ejemplos: '
                    f'{", ".join(p["valores"])})')
            else:
                que = ("identificadores" if p["identificador"]
                       else "valores distintos")
                lineas.append(
                    f'· {p["columna"]}: {p["distintos"]} {que}, y son '
                    f'EXACTAMENTE estos: {", ".join(p["valores"])}')

        if filas:
            lineas.append("Primeras filas, para que veas la forma:")
            for f in filas[:3]:
                lineas.append("   " + " | ".join(f[:len(cols)]))
        partes.append("\n".join(lineas))

    return "\n\n".join(partes)


def hay_tablas(doc_ids: list[str]) -> bool:
    if not doc_ids:
        return False
    return bool(db.uno("SELECT 1 AS x FROM app.tabla WHERE documento_id = ANY(%s) "
                       "LIMIT 1", (doc_ids,)))
